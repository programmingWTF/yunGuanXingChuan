"""
文献 Excel 导入工具 - 将整理好的文献清单导入四库中的文献库（journal_article）

用法：
    python scripts/import_xlsx_literature.py --preview                     # 只预览转换结果，不写入
    python scripts/import_xlsx_literature.py --xlsx "C:/.../万方资料整理.xlsx"   # 指定文件（可多次）
    python scripts/import_xlsx_literature.py                               # 默认导入 scripts/../data/imports/ 下所有 xlsx

说明：
- 两个整理文件（万方资料整理.xlsx / 国际传播文献整理.xlsx）列结构几乎一致：
  文献标题 / 作者 / 类型 / 摘要 / 关键词 / 主要内容 / 文章结论 / (引用文献|参考文献)
  仅最后一列列名不同（引用文献 vs 参考文献），按位置解析即可。
- 输出：data/libraries/journal_article/ 下一个 .md 一篇，格式与现有种子文档一致。
"""
import argparse
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl：pip install openpyxl")
    sys.exit(1)

PROJECT_ROOT = Path(__file__).parent.parent
IMPORT_DIR = PROJECT_ROOT / "data" / "imports"
OUTPUT_DIR = PROJECT_ROOT / "data" / "libraries" / "journal_article"

# 兼容两文件的列名（万方：引用文献 / 国际传播：参考文献）
REF_KEYS = {"引用文献", "参考文献"}


def clean_author(author: str) -> str:
    """清洗作者：数字上标转为顿号分隔（如 '庞华1周子涵1叶建宏2' -> '庞华、周子涵、叶建宏'）"""
    if not author:
        return ""
    s = str(author).strip()
    # 数字/上标替换为顿号：'庞华1周子涵1' -> '庞华、周子涵、'
    s = re.sub(r"[0-9①②③④⑤⑥⑦⑧⑨⑩]+", "、", s)
    # 已有分隔符归一
    s = re.sub(r"[,，;；\s]+", "、", s)
    s = re.sub(r"、+", "、", s).strip("、")
    return s


def clean_text(text) -> str:
    """清洗文本：去掉多余空白，压缩连续换行为单换行"""
    if not text:
        return ""
    s = str(text).strip()
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n\s*\n+", "\n", s)
    return s


def parse_workbook(path: Path) -> list:
    """解析单个 xlsx，返回文档列表"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h else "" for h in rows[0]]

    # 定位列（兼容列名差异）
    col_idx = {}
    for i, h in enumerate(header):
        if h == "文献标题":
            col_idx["title"] = i
        elif h == "作者":
            col_idx["author"] = i
        elif h == "类型":
            col_idx["type"] = i
        elif h == "摘要":
            col_idx["abstract"] = i
        elif h == "关键词":
            col_idx["keywords"] = i
        elif h == "主要内容":
            col_idx["content"] = i
        elif h == "文章结论":
            col_idx["conclusion"] = i
        elif h in REF_KEYS:
            col_idx["references"] = i

    required = {"title", "author", "abstract", "content", "conclusion"}
    missing = required - set(col_idx.keys())
    if missing:
        print(f"  ⚠ {path.name}: 缺少必需列 {missing}，跳过")
        return []

    docs = []
    for r in rows[1:]:
        if not r or not r[col_idx["title"]]:
            continue
        title = clean_text(r[col_idx["title"]])
        if not title:
            continue
        doc = {
            "title": title,
            "author": clean_author(r[col_idx.get("author", 1)]),
            "type": clean_text(r[col_idx.get("type", 2)]),
            "abstract": clean_text(r[col_idx.get("abstract", 3)]),
            "keywords": clean_text(r[col_idx.get("keywords", 4)]),
            "content": clean_text(r[col_idx.get("content", 5)]),
            "conclusion": clean_text(r[col_idx.get("conclusion", 6)]),
            "references": clean_text(r[col_idx["references"]]) if "references" in col_idx and col_idx["references"] < len(r) else "",
        }
        docs.append(doc)
    return docs


def render_markdown(doc: dict, idx: int) -> str:
    """渲染单篇文献为 markdown（格式与现有种子文档一致）"""
    lines = []
    lines.append(f"# {doc['title']}")
    lines.append("")
    if doc["author"]:
        lines.append(f"**作者**：{doc['author']}")
    if doc["type"]:
        lines.append(f"**类型**：{doc['type']}")
    if doc["keywords"]:
        lines.append(f"**关键词**：{doc['keywords']}")
    lines.append("")
    lines.append("## 摘要")
    lines.append("")
    lines.append(doc["abstract"] or "（无摘要）")
    lines.append("")
    lines.append("## 主要内容")
    lines.append("")
    lines.append(doc["content"] or "（无内容摘要）")
    lines.append("")
    lines.append("## 文章结论")
    lines.append("")
    lines.append(doc["conclusion"] or "（无结论摘要）")
    if doc["references"]:
        lines.append("")
        lines.append("## 参考文献")
        lines.append("")
        lines.append(doc["references"])
    lines.append("")
    return "\n".join(lines)


def safe_filename(title: str, idx: int) -> str:
    """生成安全文件名：清洗非法字符 + 序号防重名"""
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "", title)
    cleaned = cleaned.strip(" ._")[:50] or "untitled"
    return f"{idx:03d}_{cleaned}.md"


def main():
    parser = argparse.ArgumentParser(description="文献 Excel 导入文献库")
    parser.add_argument("--xlsx", action="append", default=[], help="指定 xlsx 文件路径（可多次）")
    parser.add_argument("--preview", action="store_true", help="只预览第一篇转换结果")
    parser.add_argument("--output", type=str, default=str(OUTPUT_DIR), help="输出目录")
    args = parser.parse_args()

    if args.xlsx:
        paths = [Path(p) for p in args.xlsx]
    else:
        IMPORT_DIR.mkdir(parents=True, exist_ok=True)
        paths = sorted(IMPORT_DIR.glob("*.xlsx"))
        if not paths:
            print(f"未指定 --xlsx，且 {IMPORT_DIR} 下无 xlsx 文件")
            print("提示：把 xlsx 放到 data/imports/ 或使用 --xlsx 指定路径")
            return

    all_docs = []
    for p in paths:
        print(f"解析 {p.name} ...")
        docs = parse_workbook(p)
        print(f"  → {len(docs)} 篇")
        all_docs.extend(docs)

    # 去重（按标题）
    seen = set()
    unique = []
    for d in all_docs:
        if d["title"] not in seen:
            seen.add(d["title"])
            unique.append(d)
    print(f"合计 {len(all_docs)} 篇，去重后 {len(unique)} 篇")

    if args.preview and unique:
        print("\n" + "=" * 60)
        print("预览第一篇：")
        print(render_markdown(unique[0], 1)[:1200])
        return

    out_dir = Path(args.output)
    out_dir.mkdir(parents=True, exist_ok=True)
    written = 0
    for i, doc in enumerate(unique, 1):
        fname = safe_filename(doc["title"], i)
        fpath = out_dir / fname
        fpath.write_text(render_markdown(doc, i), encoding="utf-8")
        written += 1
    print(f"\n✓ 已写入 {written} 篇到 {out_dir}")


if __name__ == "__main__":
    main()
